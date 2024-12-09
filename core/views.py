from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Q
from django.http import HttpResponse
from django.core.mail import send_mail
from .models import Nurse, NurseShift, Contract
from .forms import CostCalculatorForm, AdminSearchForm, ContractForm, ContractFilterForm, NurseShiftForm
import openpyxl
from datetime import datetime

def register_nurse_view(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        phone = request.POST.get('phone')
        email = request.POST.get('email')
        city = request.POST.get('city')
        specialization = request.POST.get('specialization', '')
        accepted = request.POST.get('accepted_gdpr', 'off') == 'on'

        cv_file = request.FILES.get('cv_file')
        license_file = request.FILES.get('license_file')

        if not accepted:
            return render(request, 'core/register_nurse.html', {
                'error_message': 'Du måste godkänna GDPR för att fortsätta.'
            })

        nurse = Nurse(
            name=name,
            phone=phone,
            email=email,
            city=city,
            specialization=specialization,
            accepted_gdpr=accepted
        )
        nurse.save()

        if cv_file:
            nurse.cv_file = cv_file
        if license_file:
            nurse.license_file = license_file
        nurse.save()

        subject = "Ny sjuksköterska registrerad"
        message = f"En ny sjuksköterska har registrerats:\n\nNamn: {name}\nE-post: {email}\nStad: {city}\nSpecialisering: {specialization}"
        send_mail(subject, message, 'noreply@bidma.se', ['rekrytering@bidma.se'], fail_silently=True)

        return redirect('thank_you')

    return render(request, 'core/register_nurse.html')

def thank_you_view(request):
    return render(request, 'core/thank_you.html')

@staff_member_required
def admin_search_view(request):
    q = request.GET.get('q', '')
    specialization_filter = request.GET.get('specialization', '')
    city_filter = request.GET.get('city', '')

    nurses = Nurse.objects.all().order_by('name')

    if q:
        nurses = nurses.filter(
            Q(name__icontains=q) |
            Q(email__icontains=q) |
            Q(city__icontains=q) |
            Q(specialization__icontains=q)
        )

    if specialization_filter:
        nurses = nurses.filter(specialization=specialization_filter)
    if city_filter:
        nurses = nurses.filter(city=city_filter)

    filter_form = AdminSearchForm(initial={
        'specialization': specialization_filter,
        'city': city_filter
    })

    return render(request, 'core/admin_search.html', {
        'query': q,
        'nurses': nurses,
        'filter_form': filter_form,
    })

@staff_member_required
def export_nurses_excel_view(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sjuksköterskor"

    headers = ["Namn", "E-post", "Stad", "Specialisering", "Registrerad"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    nurses = Nurse.objects.all().order_by('created_at')
    row_num = 2
    for n in nurses:
        ws.cell(row=row_num, column=1, value=n.name)
        ws.cell(row=row_num, column=2, value=n.email)
        ws.cell(row=row_num, column=3, value=n.city)
        ws.cell(row=row_num, column=4, value=n.specialization)
        ws.cell(row=row_num, column=5, value=n.created_at.strftime("%Y-%m-%d %H:%M"))
        row_num += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="nurses.xlsx"'
    wb.save(response)
    return response

@staff_member_required
def cost_calculator_view(request):
    # Samma kod som i föregående stabila version för kostnadskalkylatorn
    monthly_threshold = 46437
    low_pension_percent = 4.5
    high_pension_percent = 30.0
    sick_percent = 10.0

    base_rates = {
        ('VANLIG',1):573,('VANLIG',2):614,('VANLIG',3):665,
        ('ANESTESI',1):716,('ANESTESI',2):767,('ANESTESI',3):818,
        ('INTENSIV',1):716,('INTENSIV',2):767,('INTENSIV',3):818,
        ('OPERATION',1):716,('OPERATION',2):767,('OPERATION',3):818,
        ('BARNMORSKA',1):716,('BARNMORSKA',2):767,('BARNMORSKA',3):818,
        ('BARN',1):665,('BARN',2):716,('BARN',3):767,
        ('DISTRIKT',1):716,('DISTRIKT',2):767,('DISTRIKT',3):818,
        ('PSYKIATRI',1):665,('PSYKIATRI',2):716,('PSYKIATRI',3):767,
        ('ONKOLOGI',1):665,('ONKOLOGI',2):716,('ONKOLOGI',3):767,
        ('GERIATRIK',1):665,('GERIATRIK',2):716,('GERIATRIK',3):767,
        ('RÖNTGEN',1):573,('RÖNTGEN',2):614,('RÖNTGEN',3):665,
        ('MOTTAGNING',1):665,('MOTTAGNING',2):716,('MOTTAGNING',3):767,
        ('SKOL',1):665,('SKOL',2):716,('SKOL',3):767,
        ('INFEKTION',1):665,('INFEKTION',2):716,('INFEKTION',3):767,
    }

    add_evening = 35
    add_night = 76
    add_weekend_evening = 89
    add_weekend_day = 89
    add_weekend_night = 102
    add_storhelg_day = 171
    add_storhelg_night = 206

    jour_factor_weekend = 0.60
    beredskap_factor_weekend = 0.30
    jour_factor_weekday = 0.30
    beredskap_factor_weekday = 0.14
    jour_beredskap_aktiv_factor = 2.0

    if request.method == 'POST':
        form = CostCalculatorForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            spec = data['specialization']
            zon = int(data['zon'])
            nurse_wage = data['nurse_wage']

            regular_hours = data['regular_hours']
            evening_hours = data['evening_hours']
            night_hours = data['night_hours']
            weekend_evening_hours = data['weekend_evening_hours']
            weekend_day_hours = data['weekend_day_hours']
            weekend_night_hours = data['weekend_night_hours']
            storhelg_day_hours = data['storhelg_day_hours']
            storhelg_night_hours = data['storhelg_night_hours']

            jour_hours_weekend = data['jour_hours_weekend']
            beredskap_hours_weekend = data['beredskap_hours_weekend']
            jour_hours_weekday = data['jour_hours_weekday']
            beredskap_hours_weekday = data['beredskap_hours_weekday']
            jour_beredskap_aktiv_hours = data['jour_beredskap_aktiv_hours']

            sick_days_15_90 = data['sick_days_15_90']
            sick_days_15_45 = data['sick_days_15_45_if_under_year']
            social_fee_percent = data['social_fee_percent']

            base_rate = base_rates.get((spec, zon), 0)

            invoice_base = base_rate * regular_hours
            invoice_add = (
                evening_hours * add_evening +
                night_hours * add_night +
                weekend_evening_hours * add_weekend_evening +
                weekend_day_hours * add_weekend_day +
                weekend_night_hours * add_weekend_night +
                storhelg_day_hours * add_storhelg_day +
                storhelg_night_hours * add_storhelg_night
            )

            jour_weekend_pay = jour_hours_weekend * base_rate * jour_factor_weekend
            beredskap_weekend_pay = beredskap_hours_weekend * base_rate * beredskap_factor_weekend
            jour_weekday_pay = jour_hours_weekday * base_rate * jour_factor_weekday
            beredskap_weekday_pay = beredskap_hours_weekday * base_rate * beredskap_factor_weekday
            jour_beredskap_aktiv_pay = jour_beredskap_aktiv_hours * base_rate * jour_beredskap_aktiv_factor

            total_jour_beredskap_pay = (jour_weekend_pay + beredskap_weekend_pay +
                                        jour_weekday_pay + beredskap_weekday_pay +
                                        jour_beredskap_aktiv_pay)

            total_hours = (regular_hours + evening_hours + night_hours +
                           weekend_evening_hours + weekend_day_hours + weekend_night_hours +
                           storhelg_day_hours + storhelg_night_hours +
                           jour_hours_weekend + beredskap_hours_weekend +
                           jour_hours_weekday + beredskap_hours_weekday +
                           jour_beredskap_aktiv_hours)

            month_salary = nurse_wage * total_hours

            if month_salary <= monthly_threshold:
                pension = month_salary * (low_pension_percent/100.0)
            else:
                pension = (monthly_threshold * (low_pension_percent/100.0)) + ((month_salary - monthly_threshold)*(high_pension_percent/100.0))

            daily_salary = month_salary/30.0 if month_salary>0 else 0
            sick_pay_add = ((sick_days_15_90 * daily_salary * (sick_percent/100.0)) +
                            (sick_days_15_45 * daily_salary * (sick_percent/100.0)))

            revenue = invoice_base + invoice_add + total_jour_beredskap_pay
            social_fees = (month_salary + sick_pay_add)*(social_fee_percent/100.0)
            total_cost = month_salary + sick_pay_add + pension + social_fees
            margin = revenue - total_cost
            margin_percent = (margin/revenue)*100 if revenue > 0 else 0

            context = {
                'form': form,
                'result': True,
                'base_pay': invoice_base,
                'add_pay': invoice_add,
                'total_jour_beredskap_pay': total_jour_beredskap_pay,
                'revenue': revenue,
                'month_salary': month_salary,
                'sick_pay_add': sick_pay_add,
                'pension': pension,
                'social_fees': social_fees,
                'total_cost': total_cost,
                'margin': margin,
                'margin_percent': margin_percent,
                'total_hours': total_hours,
                'nurse_wage': nurse_wage,
            }
            return render(request, 'admin/cost_calculator.html', context)
    else:
        form = CostCalculatorForm()

    return render(request, 'admin/cost_calculator.html', {'form': form})

@staff_member_required
def admin_calendar_view(request):
    return render(request, 'admin/calendar.html')

@staff_member_required
def admin_calendar_events(request):
    from django.http import JsonResponse
    shifts = NurseShift.objects.select_related('nurse').all()
    events = []
    for s in shifts:
        events.append({
            'title': s.nurse.name,
            'start': s.start_time.isoformat(),
            'end': s.end_time.isoformat(),
            'url': f"/admin_shift_detail/{s.id}/"
        })
    return JsonResponse(events, safe=False)

@staff_member_required
def admin_add_shift_view(request):
    if request.method == 'POST':
        form = NurseShiftForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('admin_calendar')
    else:
        form = NurseShiftForm()

    return render(request, 'admin/add_shift.html', {'form': form})

@staff_member_required
def admin_nurse_detail_view(request, nurse_id):
    nurse = get_object_or_404(Nurse, id=nurse_id)
    return render(request, 'admin/nurse_detail.html', {'nurse': nurse})

@staff_member_required
def admin_shift_detail_view(request, shift_id):
    shift = get_object_or_404(NurseShift, id=shift_id)
    return render(request, 'admin/shift_detail.html', {'shift': shift})

@staff_member_required
def admin_shift_delete_view(request, shift_id):
    shift = get_object_or_404(NurseShift, id=shift_id)
    if request.method == 'POST':
        shift.delete()
        return redirect('admin_calendar')
    return render(request, 'admin/shift_delete_confirm.html', {'shift': shift})

@staff_member_required
def admin_contracts_view(request):
    form = ContractFilterForm(request.GET or None)
    contracts = Contract.objects.all().order_by('-uploaded_at')

    if form.is_valid():
        category = form.cleaned_data.get('category', '')
        name = form.cleaned_data.get('name', '')
        if category:
            contracts = contracts.filter(category=category)
        if name:
            contracts = contracts.filter(name__icontains=name)

    return render(request, 'admin/admin_contracts.html', {
        'form': form,
        'contracts': contracts,
    })

@staff_member_required
def admin_contracts_add_view(request):
    if request.method == 'POST':
        form = ContractForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('admin_contracts')
    else:
        form = ContractForm()
    return render(request, 'admin/admin_contracts_add.html', {'form': form})
